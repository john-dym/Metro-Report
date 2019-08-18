"""
Python 3.7
Targeted OS: Windows
Status: Runs
"""

#Third Party Python Libraries (xlrd, openpyxl)
from xlrd import open_workbook, XLRDError
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment

#Standard Python Libraries
from tkinter import filedialog, messagebox
from os.path import getmtime
from os import system
from time import localtime, asctime
import sys, zipfile, configparser
from ini_init import iniInit

# Doors, column labels, and metro number variables
met = []
door = []
xlCol = ['A','B','C','D','E','F']
cPartNo = 'Part No'
cPartName = 'Part Name'
cEO = 'EO No'
cLot = 'LOT No'
cLoc = 'Loc. No'
cCaseNo = 'Case No'

#Checks if the metro.ini exists and not empty. After creating the file it will attempt to open notepad to edit.
config=configparser.ConfigParser()
if config.read('metro.ini') == []:
    iniInit()
    messagebox.showerror('Error', 'Creating metro.ini. Please edit the file, save and restart the program')
    system('start ' + 'metro.ini')
    sys.exit()

#Imports metro.ini file and populates lists.
config.read('metro.ini')

for i in config['Door = Metro No']:
    door.append(i.upper())
    met.append(config['Door = Metro No'][i])

#Opens a file dialog and sets the variable wbFile with the path name
wbFile = filedialog.askopenfilename()
if wbFile == '':
    messagebox.showerror('Error','No file selected. Ending program.')
    sys.exit()

#Attempts to open the selected Excel file. If the file is not an Excel file or encrpyted the program will end.
while True:
    try:
        wb = open_workbook(wbFile)
        break
    except XLRDError:
        messagebox.showerror('Error', 'Unsupported format, or encrypted file. Try unencrypting in next window then reselect file')
        wbFile = filedialog.askopenfilename()
    except:
        messagebox.showerror('Error', 'Unknown error. Closing Program')
        sys.exit()

ws = wb.sheet_by_index(0)

# Finds the column number for Part No, Part Name, EO, Lot, Loc, Case No.
for i in range(ws.ncols):
    if ws.cell(0, i).value == cPartNo:
        cPartNo = int(i)
        break
for i in range(ws.ncols):
    if ws.cell(0, i).value == cPartName:
        cPartName = int(i)
        break
for i in range(ws.ncols):
    if ws.cell(0, i).value == cEO:
        cEO = int(i)
        break
for i in range(ws.ncols):
    if ws.cell(0, i).value == cLot:
        cLot = int(i)
        break
for i in range(ws.ncols):
    if ws.cell(0, i).value == cLoc:
        cLoc = int(i)
        break

#Defines the list variables
combPartNos = []
combPartNames = []
combPartLocs = []
combPartQtys = []
combPartEOs = []
combPartDoors = []

# Iterates through the rows to find unique parts for matching locations
for x in range(len(met)):
#Declares and erases the lists
    partNos = []
    partNames = []
    partLocs = []
    partQtys = []
    partEOs = []
    partDoors = []
    for i in range(ws.nrows):
        if ws.cell(i, cLoc).value == met[x]:
         #Increases the box count for the part
            if ws.cell(i, cPartNo).value in partNos:
                partQtys[partNos.index(ws.cell(i, cPartNo).value)] = partQtys[partNos.index(ws.cell(i, cPartNo).value)] + 1
         #Adds new part to the list and sets the box count to 1 and EO count to 0
            else:
                partNames.append(ws.cell(i, cPartName).value)
                partNos.append(ws.cell(i, cPartNo).value)
                partLocs.append(ws.cell(i, cLoc).value)
                partQtys.append(1)
                partEOs.append(0)
                partDoors.append(door[x])
        #Checks to see if the EO or Lot cell is blank, If not blank it adds 1 to the EO/Lot Qty
            if (ws.cell(i, cEO).value != '' or ws.cell(i, cLot).value != ''):
                partEOs[partNos.index(ws.cell(i, cPartNo).value)] = partEOs[partNos.index(ws.cell(i, cPartNo).value)] + 1
    #Combines lists from previous loops
    combPartNos = combPartNos + partNos
    combPartNames = combPartNames + partNames
    combPartLocs = combPartLocs + partLocs
    combPartQtys = combPartQtys + partQtys
    combPartEOs = combPartEOs + partEOs
    combPartDoors = combPartDoors + partDoors

wb = Workbook()
ws = wb.active
thin = Side(border_style='thin', color='000000') #Creates a variable for setting borders

#Adds title with date stamp and column labels to final report
ws['A1'] = 'Metro Box List         ' + str(asctime(localtime(getmtime(wbFile)))) #Writes the date/timestamp of the file exported from database in G1
ws['A1'].alignment = Alignment(vertical='center',horizontal='center')
ws.merge_cells('A1:F1')
ws['A2'] = 'Part Number'
ws['B2'] = 'Part Name'
ws['C2'] = 'Boxes'
ws['D2'] = 'EO/Lot Qty'
ws['E2'] = 'Loc. No'
ws['F2'] = 'Door'

#Shades the column headers
for x in range(ws.max_column):
    for i in range(ws.max_row):
        ws[xlCol[x]+str(i+1)].fill = PatternFill(start_color='e6e6e6', fill_type='solid')

#Fills the excel table with data
for i in range(3,len(combPartNos)):
    ws['A'+str(i)] = combPartNos[i - 3]
    ws['B'+str(i)] = combPartNames[i - 3]
    ws['C'+str(i)] = combPartQtys[i - 3]
    if combPartEOs[i - 3] != 0:         #Ignores writing a 0 in the cell
        ws['D'+str(i)] = combPartEOs[i - 3]
        ws['D'+str(i)].fill = PatternFill(start_color='ffff00', fill_type='solid')
    ws['E'+str(i)] = combPartLocs[i - 3]
    ws['F'+str(i)] = combPartDoors[i - 3]

#Apply Borders to all cells
for x in range(ws.max_column):
    for i in range(ws.max_row):
        ws[xlCol[x]+str(i+1)].border = Border(top=thin, bottom=thin, left=thin, right=thin)
        
#Preps the Excel report to fit page with minimum margins
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 6
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10

ws.print_options.horizontalCentered = True
ws.print_options.verticalCentered = True
ws.page_margins.top = 0.3
ws.page_margins.bottom = 0.3
ws.page_margins.left = 0.3
ws.page_margins.right = 0.3
ws.page_margins.header = 0
ws.page_margins.footer = 0

#Saves the finished report as tmp.xlsx
while True:
    try:
        wb.save(filename = 'tmp.xlsx')
        break
    except PermissionError:
        if messagebox.askretrycancel('Error', 'Permission Error: Is the file open? Close and retry') == False:
            sys.exit()
    except:
        messagebox.showerror('Error', 'Unknown error. Closing Program')
        sys.exit()

system('start ' + 'tmp.xlsx')
